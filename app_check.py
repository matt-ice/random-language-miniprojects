import bs4
import urllib.request
from openpyxl import Workbook
from openpyxl import load_workbook

file = "H:\classify.xlsx"
wb = load_workbook(file, data_only=True)
ws = wb["Sheet1"]

def get_res(search):
	url  = "https://www.google.com/search?q=" + str(search).replace(" ","+")
	req  = urllib.request.Request(url,headers={'User-Agent': 'Mozilla/5.0'})
	html = urllib.request.urlopen(req).read()
	soup = bs4.BeautifulSoup(html, "html.parser")
	res  = soup.find("div", {"id" : "ires"}).find("h3", {"class" : "r"}).find("a").get_text()
	return str(res)

def get_play(s):
	return get_res(s + "google play")

def get_iOS(s):
	return get_res(s + "app store")

for i in range(100, 2883):
#for i in range(31, 40):
	cell = ws["A"+str(i)].value
	try:
		ws["B"+str(i)] = get_res(cell)
	except:
		ws["B"+str(i)] = "Fail"
	try:
		ws["C"+str(i)] = get_play(cell)
	except:
		ws["C"+str(i)] = "Fail"
	try:
		ws["D"+str(i)] = get_iOS(cell)
	except:
		ws["D"+str(i)] = "Fail"
	if(i % 10 ==0):
		print(i + " passed")
	if(i%500 == 0):
		wb.save(file)
wb.save(file)


#https://play.google.com/store/search?q=test
